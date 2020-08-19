import random
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color, NamedStyle

def coloredStyle(title_color, header_color, body_color):

    def setStyle(function):

        def wrapper(ws, number_of_row):

            # 设置标题样式
            ws.merge_cells('a1:f1')
            color1 = Color(rgb=title_color)
            font = Font(name="Microsoft YaHei UI", size=34, b=True, color=color1)
            a1 = ws['a1']
            alignment = Alignment(horizontal='center', vertical='center')
            a1.font = font
            a1.alignment = alignment

            # 设置表头样式
            color2 = Color(rgb=header_color)
            style_for_row2 = NamedStyle(name='header')
            style_for_row2.font = Font(name='Calibri', size=16, color='FFFFFF')
            style_for_row2.alignment = Alignment(horizontal='center', vertical='center')
            style_for_row2.fill = PatternFill('solid', fgColor=color2)

            for each_cell in ws[2]:
                each_cell.style = style_for_row2

            # 设置表格样式
            for i in range(1, number_of_row + 3):
                ws.row_dimensions[i].height = 49.5
            for i in range(1, 7):
                ws.column_dimensions[chr(64 + i)].width = 15

            color3 = Color(rgb=body_color)
            style_for_body = NamedStyle(name='body')
            style_for_body.font = Font(name='Calibri')
            style_for_body.alignment = Alignment(horizontal='center', vertical='center')
            style_for_body.fill = PatternFill("solid", fgColor=color3)
            style_for_body.border = Border(left=Side(border_style='thin', color='FF000000'),
                                           right=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'),
                                           top=Side(border_style='thin', color='FF000000')
                                           )
            for i in range(3, number_of_row + 3):
                for j in range(1, 7):
                    ws.cell(row=i, column=j).style = style_for_body

            return function(ws, number_of_row)

        return wrapper

    return setStyle

@coloredStyle('d9530a', 'ec9800', 'e36d4a')
def sales_list(ws, number_of_row):
    ws['a1'].value = "销售清单"
    ws.cell(row=2, column=1, value="产品ID")
    ws.cell(row=2, column=2, value="名称")
    ws.cell(row=2, column=3, value="描述")
    ws.cell(row=2, column=4, value="单价")
    ws.cell(row=2, column=5, value="销售数量")
    ws.cell(row=2, column=6, value="销售金额")

    for i in range(3, number_of_row + 3):
        ws.cell(row=i, column=1, value="IN" + str(i).zfill(4))
        ws.cell(row=i, column=2, value='项目' + str(i))
        ws.cell(row=i, column=3, value='描述' + str(i))
        ws.cell(row=i, column=4, value=random.randrange(14, 60))
        ws.cell(row=i, column=5, value=random.randrange(5, 200))
        ws.cell(row=i, column=6, value=ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value)

    return ws


wb = openpyxl.Workbook()
ws = wb.active

sales_list(ws, 20).parent.save("new.xlsx")
