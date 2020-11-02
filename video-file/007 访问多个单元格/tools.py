import os
import openpyxl
import random
import decimal
import re
import configparser


# 返回一个目录内所有的.xlsx文件的路径
def excel_file_list(path: str) -> list:
    result_list = []

    for each_entry in os.scandir(path):
        # 如果文件记录是一个文件夹/目录,
        if each_entry.is_dir():
            result_list += excel_file_list(each_entry.path)

        else:
            if each_entry.path.endswith('.xlsx'):
                result_list.append(each_entry.path)

    return result_list


# 返回一个目录内所有的.xlsx文件的路径
def excel_file_list_iter(path: str) -> list:
    result_list = []

    stack = [path]

    while len(stack) != 0:
        current_dir = stack.pop()
        if os.path.isdir(current_dir):
            for each_entry in os.scandir(current_dir):
                if each_entry.is_dir():
                    stack.append(each_entry.path)
                else:
                    if each_entry.path.endswith('.xlsx'):
                        result_list.append(each_entry.path)

        else:
            if current_dir.path.endswith('.xlsx'):
                result_list.append(current_dir.path)

    return result_list


# 月度化一个工作簿
def monthlize(path: str):
    wb = openpyxl.load_workbook(path)

    for i in range(12):
        ws = wb.copy_worksheet(wb.active)
        ws.title = "{}月{}".format(i + 1, wb.active.title)
        ws.sheet_properties.tabColor = '{0:06X}'.format(random.randint(0, 0xFFFFFF))

    wb.remove(wb.active)
    wb.save(path.split('.')[0] + 'monthly.xlsx')


# 将读出的单元格转换成定点数
def transfer_to_decimal(num) -> decimal.Decimal:
    # 如果是一个整数, 就直接进行转换
    if type(num) == int:
        return decimal.Decimal(num)

    # 如果是一个浮点数, 将其转换成2位小数的字符串表示, 然后使用字符串来创建Decimal对象
    elif type(num) == float:
        return decimal.Decimal('{0:.2f}'.format(num))

    # 如果是一个字符串, 需要判断其格式
    elif type(num) == str:

        # 判断字符串是不是一个十进制的小数的表示
        if re.match('[+-]?\\d+(\\.\\d+)?$', num):
            # 字符串是一个十进制的小数表示
            # 判断是否是整数, 如果是整数, 直接通过整数创建Decimal对象
            if num.find('.') == -1:
                return decimal.Decimal(num)

            # 不是整数的情况下
            else:
                split_num = num.split('.')

                if len(split_num[1]) > 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    if int(split_num[1][2]) >= 5:
                        if num[0] == '-':
                            return decimal.Decimal(num_string) - decimal.Decimal('0.01')
                        else:
                            return decimal.Decimal(num_string) + decimal.Decimal('0.01')
                    else:
                        return decimal.Decimal(num_string)

                elif len(split_num[1]) == 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    return decimal.Decimal(num_string)

                elif len(split_num[1]) == 1:
                    num_string = split_num[0] + '.' + split_num[1] + '0'
                    return decimal.Decimal(num_string)

                else:
                    raise AttributeError

        else:
            raise AttributeError

    # 不是上述三种类型
    else:
        raise AttributeError


# 打开文件并返回指定名称的工作表, 或者返回活动工作表
def open_xlsx_file(file_name: str, sheet_name=None):
    if sheet_name:
        return openpyxl.load_workbook(file_name)[sheet_name]
    else:
        return openpyxl.load_workbook(file_name).active


# 填充金蝶导出的凭证的一列
def fill_column(col_number: int, worksheet):
    """
    处理金蝶导出的凭证, 自动填充指定的列缺失的内容
    :param col_number: 要填充的列号
    :param worksheet: 要填充的工作表对象
    :return: 填充完成后的同一个工作表对象
    """

    # 第一行固定指向2
    current = 2
    # 获取最大行
    max_index = worksheet.max_row - 1
    # While current不越界:
    while current <= max_index:
        # if格子是None
        if not worksheet.cell(row=current, column=col_number).value:
            # 填充上一格数据
            worksheet.cell(current, col_number, worksheet.cell(current - 1, col_number).value)
        # else:
        # 什么也不做

        # current移动1格
        current += 1

    return worksheet


# 加载配置文件
def load_config(file_name: str = None) -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    if not file_name:
        config.read('config.ini')
    else:
        config.read(file_name)

    return config


# 计算流动比率
def cal_current_ratio(worksheet, config) -> str:
    total_liquid_asset = worksheet[config['ratio']['liquid_asset']].value

    total_liquid_liability = worksheet[config['ratio']['liquid_liability']].value

    return "{:.2f}%".format(total_liquid_asset / total_liquid_liability * 100)

# 分列函数
def text_to_columns(index: int, delimiter: str, number: int, worksheet):
    # 首先确定填充范围, 也是从2-max_rows-1
    max_row = ws.max_row

    # 然后需要准备空行, 注意将结果分割成number列, 即在index列后插入number-1列, 所以先插入列
    worksheet.insert_cols(index + 1, number - 1)

    # 然后把列头填充一下, 从index 填充到index+number-1, 一共 number列, 这里因为是range所以正好替我们减了1
    for i in range(index, index + number):
        ws.cell(row=1, column=i, value=str(i - index + 1) + '级科目')

    # 之后每一行进行分割, 然后向同一行填充, 这里注意如果分割数量多, 需要控制不要填充出number的范围
    for i in range(2, max_row):
        # 分割index列的单元格的字符串
        split_cell = ws.cell(row=i, column=index).value.split(delimiter)

        # 如果分割的长度大于number, 只填充到number为止
        if len(split_cell) > number:
            # 从(i, index+j)的单元格开始横向填充number数量的三个单元格
            for j in range(0, number):
                ws.cell(row=i, column=index + j, value=split_cell[j].strip())
        # 如果分割的长度小于等于number, 就直接填充即可
        else:
            j = 0
            for each_content in split_cell:
                ws.cell(row=i, column=index + j, value=each_content.strip())
                j = j + 1

    return worksheet


# 固定处理序时账的函数
def process_worksheet(worksheet):
    # 填充5-6列
    tools.fill_column(5, worksheet)
    tools.fill_column(6, worksheet)

    # 删除列, 注意先删除右边的, 这样要删除的列号不会变化
    worksheet.delete_cols(15, 18)
    worksheet.delete_cols(10, 3)
    worksheet.delete_cols(8)
    worksheet.delete_cols(1, 4)

    # 分列, 注意此时分的是第四列
    text_to_columns(4, '-', 3, worksheet)

    return worksheet

# 以一列为键, 合并汇总另外一列的值
def subtotal_single(key_column: int, value_column: int, file_name: str) -> dict:
    wb = tools.open_xlsx_file(file_name)

    result = {}

    # 2 -> wb.max_row -1 是要处理的列
    for i in range(2, wb.max_row):

        # 键是否存在于字典中
        key = wb.cell(row=i, column=key_column).value

        if key in result:

            # 存在的话, 需要更新
            result[key] = result[key] + tools.transfer_to_decimal(wb.cell(row=i, column=value_column).value)

        # 不存在的话, 直接设置
        else:
            result[key] = tools.transfer_to_decimal(wb.cell(row=i, column=value_column).value)

    return result

# 以一列为键, 合并汇总另外多列的值, 每个值是平行汇总的
def subtotal_composite(key_column: int, value_column1: int, value_column2: int, file_name: str) -> dict:
    wb = tools.open_xlsx_file(file_name)

    result = {}
    # 2 -> wb.max_row -1 是要处理的列
    for i in range(2, wb.max_row):

        # 键是否存在于字典中
        key = wb.cell(row=i, column=key_column).value

        if key in result:
            # 如果存在, 要更新两个值
            result[key][wb.cell(row=1, column=value_column1).value] = result[key][wb.cell(row=1,
                                                                                          column=value_column1).value] + tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column1).value)
            result[key][wb.cell(row=1, column=value_column2).value] = result[key][wb.cell(row=1,
                                                                                          column=value_column2).value] + tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column2).value)

        # 不存在的话, 创建键和对应的嵌套字典, 初始值是0
        else:
            result[key] = {}
            result[key][wb.cell(row=1, column=value_column1).value] = tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column1).value)
            result[key][wb.cell(row=1, column=value_column2).value] = tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column2).value)

    return result