import random
import openpyxl
import tkinter
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color, NamedStyle

def coloredStyle(title_color, header_color, body_color):

    def setStyle(function):

        def wrapper(ws, number_of_row):

            ws.merge_cells('a1:f1')
            color1 = Color(rgb=title_color)
            font = Font(name="Microsoft YaHei UI", size=34, b=True, color=color1)
            a1 = ws['a1']
            alignment = Alignment(horizontal='center', vertical='center')
            a1.font = font
            a1.alignment = alignment

            color2 = Color(rgb=header_color)
            style_for_row2 = NamedStyle(name='header')
            style_for_row2.font = Font(name='Calibri', size=16, color='FFFFFF')
            style_for_row2.alignment = Alignment(horizontal='center', vertical='center')
            style_for_row2.fill = PatternFill('solid', fgColor=color2)

            for each_cell in ws[2]:
                each_cell.style = style_for_row2

            for i in range(1, number_of_row + 3):
                ws.row_dimensions[i].height = 49.5
            for i in range(1, 7):
                ws.column_dimensions[chr(64 + i)].width = 15

            color3 = Color(rgb=body_color)
            style_for_body = NamedStyle(name='body')
            style_for_body.font = Font(name='Calibri')
            style_for_body.alignment = Alignment(horizontal='center', vertical='center')
            style_for_body.fill = PatternFill("solid", fgColor=color3)
            style_for_body.border = Border(left=Side(border_style='thin', color='FF000000'),
                                           right=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'),
                                           top=Side(border_style='thin', color='FF000000')
                                           )
            for i in range(3, number_of_row + 3):
                for j in range(1, 7):
                    ws.cell(row=i, column=j).style = style_for_body

            return function(ws, number_of_row)

        return wrapper

    return setStyle


@coloredStyle('96c8ac', '75c1dd', 'c6da87')
def sales_list(ws, number_of_row):
    ws['a1'].value = "销售清单"
    ws.cell(row=2, column=1, value="产品ID")
    ws.cell(row=2, column=2, value="名称")
    ws.cell(row=2, column=3, value="描述")
    ws.cell(row=2, column=4, value="单价")
    ws.cell(row=2, column=5, value="销售数量")
    ws.cell(row=2, column=6, value="销售金额")

    for i in range(3, number_of_row + 3):
        ws.cell(row=i, column=1, value="IN" + str(i).zfill(4))
        ws.cell(row=i, column=2, value='项目' + str(i))
        ws.cell(row=i, column=3, value='描述' + str(i))
        ws.cell(row=i, column=4, value=random.randrange(14, 60))
        ws.cell(row=i, column=5, value=random.randrange(5, 200))
        ws.cell(row=i, column=6, value=ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value)

    return ws

if __name__ == '__main__':

    # 排布窗体和控件
    main_frame = tkinter.Tk()
    main_frame.title("生成excel文件")
    main_frame.geometry("800x600")

    main_message = '请输入要生成的文件名称, 不输入则默认保存为new.xlsx'
    message_label = tkinter.Label(master=main_frame,text=main_message, font=("微软雅黑", 24))
    message_label.grid(row=0)

    tkinter.Label(main_frame, text="文件名:", font=("微软雅黑", 16)).grid(column=0, row=1)
    input_filename = tkinter.Entry(main_frame, width=20,font=("微软雅黑", 16))
    input_filename.grid(column=0,row=2)

    start_button = tkinter.Button(main_frame, text="生成文件",font=("微软雅黑", 16),bg='#4cc7b2',fg='white')
    start_button.grid(row=3, pady=10)

    result_label = tkinter.Label(main_frame,font=("微软雅黑", 14))
    result_label.grid(row=4)

    # 创建按钮绑定的函数
    def button_on_click():
        start_button.configure(state=tkinter.DISABLED)
        filename = input_filename.get().strip() or 'new.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active

        time.sleep(2)

        try:
            sales_list(ws, 20).parent.save(filename)
            result_label.configure(text="成功保存为: {}".format(filename))
        except:
            result_label.configure(text="文件名不合法, 请重新再试".format(filename))
        finally:
            start_button.configure(state=tkinter.NORMAL)

    # 绑定函数给按钮
    start_button.configure(command=button_on_click)

    # 启动图形化界面
    main_frame.mainloop()